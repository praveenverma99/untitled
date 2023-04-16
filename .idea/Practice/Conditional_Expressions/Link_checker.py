import openpyxl
import requests

book = openpyxl.load_workbook("dev_url.xlsx")

tab1 = book.get_sheet_by_name("data1")

text_file = open("output.txt", "w")

while str(tab1.cell(row=i,column=1).value)!="None":

    x = str(tab1.cell(row=i,column=1).value)

    if x!= "None":
        r=requests.get(x)
        print(x)

    if(r.status_code==404):
        print(x)
n = text_file.write(x+"\n")

text_file.close()